from __future__ import annotations

import csv
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook


INPUT_FIELDNAMES = [
    "query",
    "rank",
    "site_name",
    "title",
    "url",
    "displayed_domain",
    "root_domain",
    "snippet",
    "provider_raw_date",
    "domain_created_at",
    "domain_created_source",
    "exclude_reason",
    "status",
]

MERGED_FIELDNAMES = [
    "best_rank",
    "query_hit_count",
    "matched_queries",
    "best_url",
    "best_title",
]

SCORE_FIELDNAMES = [
    "classification",
    "decision",
    "review_reason",
    "relevance_score",
    "quality_score",
    "final_score",
    "signal_summary",
]

REVIEW_FIELDNAMES = INPUT_FIELDNAMES + MERGED_FIELDNAMES + SCORE_FIELDNAMES

SOCIAL_VIDEO_TOKENS = ["youtube.com", "youtu.be", "x.com/", "twitter.com/", "linkedin.com/", "facebook.com/", "instagram.com/", "tiktok.com/"]
FORUM_TOKENS = ["forum.", "/forum", "community.", "/community/", "idea-exchange", "productideas", "discourse", "/threads/"]
DOC_TOKENS = ["docs.", "/docs/", "support.", "documentation", "help center", "tutorial", "learn.microsoft.com", "oracle help center"]
SERVICE_VENDOR_TOKENS = ["fiverr.com", "upwork.com", "manual submission service", "submit to 300 directories", "submit to 400 directories", "promotion platforms"]
LEGAL_OR_CONVERSION_TOKENS = ["/privacy", "privacy-policy", "/terms", "/pricing", "/login", "/signup", "/register"]
BLOG_TOKENS = ["/blog/", "/news/", "/posts/", "/article/", "/articles/", "/guides/"]
DETAIL_TOKENS = ["/tool/", "tool_details", "tool_details.php", "?slug=", "&slug=", "?id=", "&id="]
SUBMIT_TOKENS = ["/submit", "submit your tool", "submit a tool", "submit tool", "add your tool", "add tool", "list your tool", "feature your ai tool", "get listed"]
AI_TOKENS = [" ai ", "ai-", "ai/", "ai tool", "ai tools", "llm", "gpt", "artificial intelligence", "machine learning"]
DIRECTORY_TOKENS = ["directory", "catalog", "catalogue", "library", "database", "resource", "resources", "toolbox", "listing", "rankings"]
INCLUSION_TOKENS = ["feature", "list", "listing", "add listing", "get listed", "submit startup"]


@dataclass(slots=True)
class CleanedRow:
    row: dict[str, str]
    classification: str
    decision: str
    review_reason: str
    relevance_score: int
    quality_score: int
    final_score: float
    signal_summary: str

    def as_review_row(self) -> dict[str, str | int | float]:
        payload = {field: self.row.get(field, "") for field in INPUT_FIELDNAMES + MERGED_FIELDNAMES}
        payload["classification"] = self.classification
        payload["decision"] = self.decision
        payload["review_reason"] = self.review_reason
        payload["relevance_score"] = self.relevance_score
        payload["quality_score"] = self.quality_score
        payload["final_score"] = self.final_score
        payload["signal_summary"] = self.signal_summary
        return payload


def load_results_file(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["results"]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value) for value in rows[0]]
        results: list[dict[str, str]] = []
        for values in rows[1:]:
            results.append({header: "" if value is None else str(value) for header, value in zip(headers, values)})
        return results
    raise ValueError(f"Unsupported input file: {path}")


def classify_rows(rows: list[dict[str, str]]) -> list[CleanedRow]:
    return [classify_row(row) for row in rows]


def _normalized_row(row: dict[str, str]) -> dict[str, str]:
    normalized = {field: str(row.get(field, "") or "") for field in INPUT_FIELDNAMES + MERGED_FIELDNAMES}
    if not normalized["best_rank"]:
        normalized["best_rank"] = normalized.get("rank", "")
    if not normalized["query_hit_count"]:
        normalized["query_hit_count"] = "1"
    if not normalized["matched_queries"]:
        normalized["matched_queries"] = normalized.get("query", "")
    if not normalized["best_url"]:
        normalized["best_url"] = normalized.get("url", "")
    if not normalized["best_title"]:
        normalized["best_title"] = normalized.get("title", "")
    return normalized


def _haystack(row: dict[str, str]) -> str:
    return " ".join(
        [
            row.get("site_name", ""),
            row.get("title", ""),
            row.get("url", ""),
            row.get("snippet", ""),
            row.get("best_title", ""),
            row.get("best_url", ""),
        ]
    ).lower()


def _url_tokens(row: dict[str, str]) -> tuple[str, str, str]:
    url = row.get("best_url") or row.get("url", "")
    parsed = urlparse(url)
    return url.lower(), parsed.path.lower(), parsed.query.lower()


def _contains_any(text: str, tokens: list[str]) -> bool:
    return any(token in text for token in tokens)


def _parse_int(value: str) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _domain_age_years(domain_created_at: str) -> int | None:
    value = str(domain_created_at or "").strip()
    if not value:
        return None
    try:
        created = date.fromisoformat(value[:10])
    except ValueError:
        return None
    return date.today().year - created.year - ((date.today().month, date.today().day) < (created.month, created.day))


def classify_row(row: dict[str, str]) -> CleanedRow:
    normalized = _normalized_row(row)
    haystack = _haystack(normalized)
    url, path, query_string = _url_tokens(normalized)
    best_rank = _parse_int(normalized.get("best_rank", "")) or _parse_int(normalized.get("rank", "")) or 999
    query_hit_count = _parse_int(normalized.get("query_hit_count", "")) or 1
    domain_age_years = _domain_age_years(normalized.get("domain_created_at", ""))

    positive_signals: list[str] = []
    negative_signals: list[str] = []

    if _contains_any(haystack, SOCIAL_VIDEO_TOKENS):
        return _hard_drop(normalized, "video_or_social", "social or video result", ["social_or_video"])
    if _contains_any(haystack, FORUM_TOKENS):
        return _hard_drop(normalized, "forum_or_community", "forum or community discussion page", ["forum_or_community"])
    if _contains_any(haystack, DOC_TOKENS):
        return _hard_drop(normalized, "product_doc", "product documentation or help page", ["docs_or_help"])
    if _contains_any(haystack, SERVICE_VENDOR_TOKENS):
        return _hard_drop(normalized, "service_vendor", "manual submission service or reseller page", ["service_vendor"])
    if _contains_any(url, LEGAL_OR_CONVERSION_TOKENS):
        return _hard_drop(normalized, "legal_or_conversion_page", "privacy, pricing, login, or signup page", ["legal_or_conversion_page"])
    if _contains_any(url, BLOG_TOKENS):
        return _hard_drop(normalized, "blog_or_article", "blog, article, or editorial page", ["blog_or_article"])
    if _contains_any(url, DETAIL_TOKENS):
        return _hard_drop(normalized, "detail_page", "tool detail page instead of a submission page", ["detail_page"])

    has_submit = _contains_any(haystack, SUBMIT_TOKENS)
    has_ai = _contains_any(f" {haystack} ", AI_TOKENS)
    has_directory = _contains_any(haystack, DIRECTORY_TOKENS)
    has_inclusion = _contains_any(haystack, INCLUSION_TOKENS)

    relevance_score = 0
    quality_score = 0
    classification = "likely_irrelevant"

    if has_submit:
        relevance_score += 4
        positive_signals.append("submit_signal")
    if has_ai:
        relevance_score += 2
        positive_signals.append("ai_signal")
    if has_directory:
        relevance_score += 2
        positive_signals.append("directory_signal")
    if has_inclusion:
        relevance_score += 1
        positive_signals.append("inclusion_signal")
    if query_hit_count >= 2:
        relevance_score += 1
        positive_signals.append(f"query_hit_{query_hit_count}")

    if best_rank <= 5:
        quality_score += 3
        positive_signals.append("rank_top_5")
    elif best_rank <= 10:
        quality_score += 2
        positive_signals.append("rank_top_10")
    if query_hit_count >= 2:
        quality_score += 2
    if domain_age_years is not None and domain_age_years >= 3:
        quality_score += 2
        positive_signals.append(f"domain_age_{domain_age_years}y")
    if path in {"", "/"} or (has_submit and not query_string and path.count("/") <= 2):
        quality_score += 1
        positive_signals.append("clean_url")

    if _contains_any(url, BLOG_TOKENS):
        quality_score -= 4
        negative_signals.append("blog_path")
    if _contains_any(url, DETAIL_TOKENS):
        quality_score -= 4
        negative_signals.append("detail_page")
    if _contains_any(url, LEGAL_OR_CONVERSION_TOKENS):
        quality_score -= 4
        negative_signals.append("legal_or_conversion_page")
    if _contains_any(haystack, SERVICE_VENDOR_TOKENS):
        quality_score -= 5
        negative_signals.append("service_vendor")

    relevance_score = min(relevance_score, 10)
    quality_score = max(min(quality_score, 10), 0)
    final_score = round(relevance_score * 0.7 + quality_score * 0.3, 2)

    if has_submit and (has_ai or has_directory):
        classification = "submission_candidate"
    elif (path in {"", "/"} or path.count("/") <= 1) and has_ai and has_directory:
        classification = "directory_home_candidate"
    elif has_directory or has_inclusion:
        classification = "catalog_or_resource_candidate"

    if relevance_score >= 6 and final_score >= 7:
        decision = "keep"
        review_reason = "submission page with strong relevance and quality signals"
    elif relevance_score >= 4 and final_score >= 5:
        decision = "flag"
        review_reason = "candidate needs review due to mixed submission or quality signals"
    elif classification == "directory_home_candidate":
        decision = "flag"
        review_reason = "high-quality directory homepage without explicit submit signal"
    else:
        decision = "drop"
        review_reason = "insufficient relevance or quality signals"

    signal_summary = ";".join(positive_signals + negative_signals)
    return CleanedRow(
        row=normalized,
        classification=classification,
        decision=decision,
        review_reason=review_reason,
        relevance_score=relevance_score,
        quality_score=quality_score,
        final_score=final_score,
        signal_summary=signal_summary,
    )


def _hard_drop(row: dict[str, str], classification: str, review_reason: str, signals: list[str]) -> CleanedRow:
    return CleanedRow(
        row=row,
        classification=classification,
        decision="drop",
        review_reason=review_reason,
        relevance_score=0,
        quality_score=0,
        final_score=0.0,
        signal_summary=";".join(signals),
    )


def _sort_key(row: CleanedRow) -> tuple[int, float, int, str]:
    decision_order = {"keep": 0, "flag": 1, "drop": 2}
    best_rank = _parse_int(row.row.get("best_rank", "")) or _parse_int(row.row.get("rank", "")) or 999
    root_domain = row.row.get("root_domain", "")
    return (decision_order[row.decision], -row.final_score, best_rank, root_domain)


def write_clean_results(rows: list[CleanedRow], output_prefix: Path) -> tuple[Path, Path, Path, Path]:
    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    csv_path = output_prefix.with_suffix(".csv")
    xlsx_path = output_prefix.with_suffix(".xlsx")
    review_path = output_prefix.with_name(f"{output_prefix.name}.review.csv")
    manifest_path = output_prefix.with_name(f"{output_prefix.name}.manifest.json")

    sorted_rows = sorted(rows, key=_sort_key)

    kept_or_flagged: list[dict[str, str | int | float]] = []
    seen_domains: set[str] = set()
    for row in sorted_rows:
        if row.decision not in {"keep", "flag"}:
            continue
        root_domain = row.row.get("root_domain", "")
        if root_domain and root_domain in seen_domains:
            continue
        if root_domain:
            seen_domains.add(root_domain)
        kept_or_flagged.append(row.as_review_row())

    review_rows = [row.as_review_row() for row in sorted_rows]

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_FIELDNAMES)
        writer.writeheader()
        writer.writerows(kept_or_flagged)

    with review_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_FIELDNAMES)
        writer.writeheader()
        writer.writerows(review_rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(REVIEW_FIELDNAMES)
    for row in kept_or_flagged:
        sheet.append([row.get(field, "") for field in REVIEW_FIELDNAMES])
    workbook.save(xlsx_path)

    decision_counts = Counter(row.decision for row in rows)
    classification_counts = Counter(row.classification for row in rows)
    manifest = {
        "input_count": len(rows),
        "output_count": len(kept_or_flagged),
        "decision_counts": dict(decision_counts),
        "classification_counts": dict(classification_counts),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "review_csv_path": str(review_path),
        "deduped_root_domains": len(seen_domains),
        "sort_order": ["decision", "final_score_desc", "best_rank_asc"],
    }
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    return csv_path, xlsx_path, review_path, manifest_path
