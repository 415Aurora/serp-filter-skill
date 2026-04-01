from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from serp_filter.domain_utils import normalize_root_domain


AUTO_COLUMN_HINTS = ("url", "link", "site", "domain", "website", "submit")


@dataclass(slots=True)
class BlocklistSourceConfig:
    path: Path
    sheet_name: str | None = None
    url_columns: list[str] = field(default_factory=list)
    domain_columns: list[str] = field(default_factory=list)


def load_blocked_domains(config: BlocklistSourceConfig) -> set[str]:
    suffix = config.path.suffix.lower()
    if suffix == ".xlsx":
        return _load_blocked_domains_from_xlsx(config)
    if suffix in {".csv", ".txt"}:
        return _load_blocked_domains_from_text(config.path)
    raise ValueError(f"Unsupported blocklist format: {config.path.suffix}")


def _load_blocked_domains_from_xlsx(config: BlocklistSourceConfig) -> set[str]:
    workbook = load_workbook(config.path, read_only=True, data_only=True)
    sheet = workbook[config.sheet_name] if config.sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return set()

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    column_indexes = _resolve_column_indexes(header, config.url_columns + config.domain_columns)
    if not column_indexes:
        column_indexes = _auto_detect_column_indexes(header)

    domains: set[str] = set()
    for row in rows[1:]:
        for index in column_indexes:
            if index >= len(row):
                continue
            value = row[index]
            if value is None:
                continue
            domain = normalize_root_domain(str(value))
            if domain:
                domains.add(domain)
    return domains


def _resolve_column_indexes(header: list[str], selected: Iterable[str]) -> list[int]:
    lowered = {name.strip().lower() for name in selected}
    return [index for index, value in enumerate(header) if value.lower() in lowered]


def _auto_detect_column_indexes(header: list[str]) -> list[int]:
    return [
        index
        for index, value in enumerate(header)
        if any(hint in value.lower() for hint in AUTO_COLUMN_HINTS)
    ]


def _load_blocked_domains_from_text(path: Path) -> set[str]:
    domains: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        domain = normalize_root_domain(line)
        if domain:
            domains.add(domain)
    return domains
