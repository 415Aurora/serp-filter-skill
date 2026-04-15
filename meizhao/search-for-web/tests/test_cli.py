from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook

from serp_filter.cli import main


def test_cli_run_with_static_provider_writes_outputs(tmp_path: Path) -> None:
    query_file = tmp_path / "queries.txt"
    query_file.write_text("best ai directories\n", encoding="utf-8")

    provider_data = tmp_path / "provider.json"
    provider_data.write_text(
        json.dumps(
            {
                "best ai directories": [
                    {
                        "position": 1,
                        "title": "Example Result",
                        "link": "https://www.example.com/tools",
                        "displayed_link": "www.example.com",
                        "source": "Example",
                        "snippet": "example",
                    },
                    {
                        "position": 2,
                        "title": "BBC Result",
                        "link": "https://www.bbc.co.uk/news/ai",
                        "displayed_link": "www.bbc.co.uk",
                        "source": "BBC",
                        "snippet": "bbc",
                    },
                ]
            }
        ),
        encoding="utf-8",
    )
    provider_config = tmp_path / "providers.toml"
    provider_config.write_text(
        "\n".join(
            [
                "[static_json]",
                f"data_path = {json.dumps(str(provider_data))}",
                "",
            ]
        ),
        encoding="utf-8",
    )

    blocklist = tmp_path / "blocklist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sites"
    sheet.append(["name", "submit link"])
    sheet.append(["Example", "https://example.com"])
    workbook.save(blocklist)

    output_prefix = tmp_path / "output" / "run-1"

    exit_code = main(
        [
            "run",
            "--query-file",
            str(query_file),
            "--blocklist-file",
            str(blocklist),
            "--provider",
            "static-json",
            "--provider-config",
            str(provider_config),
            "--domain-date-provider",
            "noop",
            "--output-prefix",
            str(output_prefix),
        ]
    )

    assert exit_code == 0
    assert output_prefix.with_suffix(".csv").exists()
    assert output_prefix.with_suffix(".xlsx").exists()


def test_cli_clean_writes_candidate_results_and_review_outputs(tmp_path: Path) -> None:
    source_csv = tmp_path / "input.csv"
    with source_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "query",
                "rank",
                "site_name",
                "title",
                "url",
                "displayed_domain",
                "root_domain",
                "snippet",
                "provider_raw_date",
                "domain_created_at",
                "domain_created_source",
                "exclude_reason",
                "status",
                "best_rank",
                "query_hit_count",
                "matched_queries",
                "best_url",
                "best_title",
            ],
        )
        writer.writeheader()
        writer.writerows(
            [
                {
                    "query": "submit a tool",
                    "rank": 1,
                    "site_name": "AI ProvenTools",
                    "title": "Submit A Tool - AI ProvenTools",
                    "url": "https://aiproventools.com/submit-a-tool/",
                    "displayed_domain": "aiproventools.com",
                    "root_domain": "aiproventools.com",
                    "snippet": "Submit your AI tool to our directory",
                    "domain_created_at": "2018-01-01",
                    "domain_created_source": "rdap",
                    "status": "kept",
                    "best_rank": 1,
                    "query_hit_count": 2,
                    "matched_queries": "submit a tool; ai tools directory",
                    "best_url": "https://aiproventools.com/submit-a-tool/",
                    "best_title": "Submit A Tool - AI ProvenTools",
                },
                {
                    "query": "submit a tool",
                    "rank": 2,
                    "site_name": "ServiceNow",
                    "title": "Add a tool",
                    "url": "https://www.servicenow.com/docs/r/intelligent-experiences/now-assist-skill-kit/add-a-tool.html",
                    "displayed_domain": "servicenow.com",
                    "root_domain": "servicenow.com",
                    "snippet": "Select the skill you want to add a tool to",
                    "status": "kept",
                    "best_rank": 2,
                    "query_hit_count": 1,
                    "matched_queries": "submit a tool",
                    "best_url": "https://www.servicenow.com/docs/r/intelligent-experiences/now-assist-skill-kit/add-a-tool.html",
                    "best_title": "Add a tool",
                },
                {
                    "query": "submit a tool",
                    "rank": 6,
                    "site_name": "General AI Directory",
                    "title": "AI Tools Directory",
                    "url": "https://webtoolsweekly.com/",
                    "displayed_domain": "webtoolsweekly.com",
                    "root_domain": "webtoolsweekly.com",
                    "snippet": "Discover the best AI tools in our directory",
                    "domain_created_at": "2016-01-01",
                    "domain_created_source": "rdap",
                    "status": "kept",
                    "best_rank": 6,
                    "query_hit_count": 2,
                    "matched_queries": "submit a tool; ai tools directory",
                    "best_url": "https://webtoolsweekly.com/",
                    "best_title": "AI Tools Directory",
                },
            ]
        )

    output_prefix = tmp_path / "cleaned" / "round-2"

    exit_code = main(
        [
            "clean",
            "--input-file",
            str(source_csv),
            "--output-prefix",
            str(output_prefix),
        ]
    )

    assert exit_code == 0
    assert output_prefix.with_suffix(".csv").exists()
    assert output_prefix.with_suffix(".xlsx").exists()
    assert output_prefix.with_name(f"{output_prefix.name}.review.csv").exists()
    assert output_prefix.with_name(f"{output_prefix.name}.manifest.json").exists()

    cleaned_rows = output_prefix.with_suffix(".csv").read_text(encoding="utf-8")
    assert "aiproventools.com/submit-a-tool/" in cleaned_rows
    assert "webtoolsweekly.com/" in cleaned_rows
    assert "servicenow.com/docs/" not in cleaned_rows

    review_rows = output_prefix.with_name(f"{output_prefix.name}.review.csv").read_text(encoding="utf-8")
    assert "submission_candidate" in review_rows
    assert "product_doc" in review_rows
    assert "drop" in review_rows

    workbook = load_workbook(output_prefix.with_suffix(".xlsx"), read_only=True, data_only=True)
    values = list(workbook["results"].iter_rows(values_only=True))
    assert values[0][-7:] == (
        "classification",
        "decision",
        "review_reason",
        "relevance_score",
        "quality_score",
        "final_score",
        "signal_summary",
    )

    manifest = json.loads(output_prefix.with_name(f"{output_prefix.name}.manifest.json").read_text(encoding="utf-8"))
    assert manifest["input_count"] == 3
    assert manifest["output_count"] == 2
    assert manifest["decision_counts"]["keep"] == 1
    assert manifest["decision_counts"]["flag"] == 1
    assert manifest["decision_counts"]["drop"] == 1


def test_cli_run_with_query_template_file_merges_queries(tmp_path: Path) -> None:
    query_file = tmp_path / "queries.txt"
    query_file.write_text("submit your AI tool\n", encoding="utf-8")

    template_file = tmp_path / "templates.txt"
    template_file.write_text("submit your AI tool\nadd your AI tool directory\n", encoding="utf-8")

    provider_data = tmp_path / "provider.json"
    provider_data.write_text(
        json.dumps(
            {
                "submit your AI tool": [
                    {
                        "position": 1,
                        "title": "Example Result",
                        "link": "https://www.example.com/tools",
                        "displayed_link": "www.example.com",
                        "source": "Example",
                        "snippet": "example",
                    }
                ],
                "add your AI tool directory": [
                    {
                        "position": 1,
                        "title": "Second Result",
                        "link": "https://www.second.com/tools",
                        "displayed_link": "www.second.com",
                        "source": "Second",
                        "snippet": "second",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    provider_config = tmp_path / "providers.toml"
    provider_config.write_text(
        "\n".join(
            [
                "[static_json]",
                f"data_path = {json.dumps(str(provider_data))}",
                "",
            ]
        ),
        encoding="utf-8",
    )

    blocklist = tmp_path / "blocklist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sites"
    sheet.append(["name", "submit link"])
    workbook.save(blocklist)

    output_prefix = tmp_path / "output" / "run-templates"

    exit_code = main(
        [
            "run",
            "--query-file",
            str(query_file),
            "--query-template-file",
            str(template_file),
            "--blocklist-file",
            str(blocklist),
            "--provider",
            "static-json",
            "--provider-config",
            str(provider_config),
            "--domain-date-provider",
            "noop",
            "--output-prefix",
            str(output_prefix),
        ]
    )

    assert exit_code == 0
    assert output_prefix.with_name("run-templates-01.csv").exists()
    assert output_prefix.with_name("run-templates-02.csv").exists()
    assert output_prefix.with_name("run-templates-merged.csv").exists()
    assert output_prefix.with_name("run-templates-merged.xlsx").exists()

    merged_rows = list(csv.DictReader(output_prefix.with_name("run-templates-merged.csv").open(encoding="utf-8")))
    assert len(merged_rows) == 2
    assert "best_rank" in merged_rows[0]
    assert "query_hit_count" in merged_rows[0]
