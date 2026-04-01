from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from serp_filter.models import SearchResult
from serp_filter.pipeline import run_pipeline


class _FakeProvider:
    def search(self, query: str, limit: int, locale: str | None = None) -> list[SearchResult]:
        return [
            SearchResult(
                query=query,
                rank=1,
                title="Example Result",
                site_name="Example",
                url="https://www.example.com/tools",
                displayed_domain="www.example.com",
                root_domain="example.com",
                snippet="example",
                provider_raw_date=None,
            ),
            SearchResult(
                query=query,
                rank=2,
                title="BBC Result",
                site_name="BBC",
                url="https://www.bbc.co.uk/news/ai",
                displayed_domain="www.bbc.co.uk",
                root_domain="bbc.co.uk",
                snippet="bbc",
                provider_raw_date="Mar 1, 2026",
            ),
        ]


def test_run_pipeline_filters_blocked_results_and_writes_outputs(tmp_path: Path) -> None:
    output_prefix = tmp_path / "serp-results"

    results = run_pipeline(
        query="best ai directories",
        provider=_FakeProvider(),
        blocked_domains={"example.com"},
        domain_lookup=lambda domain: ("1999-08-01", "rdap") if domain == "bbc.co.uk" else ("2000-01-01", "rdap"),
        output_prefix=output_prefix,
        limit=10,
        locale="us",
    )

    assert len(results.kept_results) == 1
    assert results.kept_results[0].root_domain == "bbc.co.uk"
    assert results.excluded_results[0].exclude_reason == "blocked_domain"
    assert results.csv_path.exists()
    assert results.xlsx_path.exists()
    assert results.manifest_path.exists()

    with results.csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 1
    assert rows[0]["root_domain"] == "bbc.co.uk"
    assert rows[0]["domain_created_at"] == "1999-08-01"

    workbook = load_workbook(results.xlsx_path, read_only=True, data_only=True)
    sheet = workbook["results"]
    values = list(sheet.iter_rows(values_only=True))
    assert values[0][:5] == ("query", "rank", "site_name", "title", "url")
    assert values[1][6] == "bbc.co.uk"


def test_run_pipeline_looks_up_each_root_domain_once(tmp_path: Path) -> None:
    output_prefix = tmp_path / "dedupe-results"
    call_log: list[str] = []

    class _DuplicateProvider:
        def search(self, query: str, limit: int, locale: str | None = None) -> list[SearchResult]:
            return [
                SearchResult(
                    query=query,
                    rank=1,
                    title="Example Root",
                    site_name="Example",
                    url="https://www.example.com/tools",
                    displayed_domain="www.example.com",
                    root_domain="example.com",
                    snippet="one",
                    provider_raw_date=None,
                ),
                SearchResult(
                    query=query,
                    rank=2,
                    title="Example Blog",
                    site_name="Example Blog",
                    url="https://blog.example.com/post",
                    displayed_domain="blog.example.com",
                    root_domain="example.com",
                    snippet="two",
                    provider_raw_date=None,
                ),
            ]

    def _lookup(domain: str) -> tuple[str | None, str]:
        call_log.append(domain)
        return "2000-01-01", "rdap"

    run_pipeline(
        query="example query",
        provider=_DuplicateProvider(),
        blocked_domains=set(),
        domain_lookup=_lookup,
        output_prefix=output_prefix,
        limit=10,
        locale="us",
    )

    assert call_log == ["example.com"]
